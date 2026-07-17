import os
import json
import google.generativeai as genai
from pathlib import Path
import time
import PyPDF2
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QPushButton, QLabel, QFileDialog, QLineEdit, 
                             QTextEdit, QProgressBar, QMessageBox, QHBoxLayout)
from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl
import sys

### Ashar API Key
# API_KEY = "AIzaSyCtTxyfboDNpJAqIQg48O5l10oT9jtcbPk"

### Mine API Key
API_KEY = "AIzaSyAAkJRHi6u1nX3I8lIUXTRYMe6v-zpMXic"


class PDFCleaner:
    """Class to handle PDF cleaning operations."""
    
    def __init__(self, keywords):
        self.keywords = keywords
    
    def extract_text_from_page(self, page):
        """Extract text from a PDF page."""
        try:
            return page.extract_text().lower()
        except:
            return ""
    
    def page_contains_keywords(self, page):
        """Check if a page contains any of the specified keywords."""
        text = self.extract_text_from_page(page)
        return any(keyword.lower() in text for keyword in self.keywords)
    
    def clean_pdf(self, input_path, output_path):
        """Clean a single PDF by extracting relevant pages."""
        try:
            with open(input_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                writer = PyPDF2.PdfWriter()
                
                total_pages = len(reader.pages)
                pages_to_extract = set()
                
                # Always include first three pages
                for i in range(min(1, total_pages)):
                    pages_to_extract.add(i)
                
                # Check all pages for keywords
                for page_num in range(total_pages):
                    page = reader.pages[page_num]
                    if self.page_contains_keywords(page):
                        pages_to_extract.add(page_num)
                
                # Add selected pages to writer in order
                for page_num in sorted(pages_to_extract):
                    writer.add_page(reader.pages[page_num])
                
                # Save the new PDF
                with open(output_path, 'wb') as output_file:
                    writer.write(output_file)
                
                return len(pages_to_extract), total_pages
                
        except Exception as e:
            raise Exception(f"Error cleaning PDF: {str(e)}")


class DataExtractor:
    """Class to handle data extraction from PDFs using Gemini API."""
    
    def __init__(self, api_key, extraction_prompt_path):
        self.api_key = api_key
        genai.configure(api_key=self.api_key)
        self.model = genai.GenerativeModel('gemini-2.5-flash')
        
        # Load extraction prompt
        with open(extraction_prompt_path, "r", encoding="utf-8") as f:
            self.extraction_prompt = f.read()
    
    def extract_data_from_pdf(self, pdf_path):
        """Extract financial data from a PDF using Gemini API."""
        try:
            if not os.path.exists(pdf_path):
                return None
            
            pdf_bytes = Path(pdf_path).read_bytes()
            
            pdf_part = {
                "mime_type": "application/pdf",
                "data": pdf_bytes
            }
            # print(self.extraction_prompt)
            response = self.model.generate_content([self.extraction_prompt, pdf_part])
            response_text = response.text.strip()
            
            # Remove markdown code blocks if present
            if response_text.startswith("```json"):
                response_text = response_text[7:]
            if response_text.startswith("```"):
                response_text = response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            
            response_text = response_text.strip()
            extracted_data = json.loads(response_text)
            
            return extracted_data
            
        except json.JSONDecodeError as e:
            print(f"Error parsing JSON response: {e}")
            raise Exception(f"JSON parsing error: {e}")
        except Exception as e:
            print(f"Error extracting data: {e}")
            raise Exception(f"API extraction error: {e}")


class ExcelMerger:
    """Class to handle merging JSON data with Excel file."""
    
    def __init__(self, excel_path, json_path):
        self.excel_path = excel_path
        self.json_path = json_path
        self.sheet_name = "AV Documents_Clean"
    
    def extract_filename_from_path(self, path_string):
        """Extract filename from a full path or return the string as is."""
        if not path_string:
            return ""
        
        # Check if it looks like a path
        if '\\' in path_string or '/' in path_string:
            return os.path.basename(path_string)
        return path_string
    
    def merge_data(self):
        """Merge JSON data into Excel file."""
        try:
            # Load JSON data
            with open(self.json_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            # Load Excel workbook
            wb = openpyxl.load_workbook(self.excel_path)
            
            if self.sheet_name not in wb.sheetnames:
                raise Exception(f"Sheet '{self.sheet_name}' not found in Excel file")
            
            ws = wb[self.sheet_name]
            
            # Find column indices
            headers = {}
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value:
                    headers[cell.value] = col_idx
            
            # Define column mapping
            column_mapping = {
                '2024 AV Link': None,
                'AV Date': None,
                'AV Date Reference': None,
                'Inflation Rate': None,
                'Inflation Rate Reference': None,
                'Rate of Return on Pension Investments': None,
                'Rate of Return Reference': None,
                'Smoothing': None,
                'Smoothing Reference': None
            }
            
            # Map columns
            av_date_col = None
            inflation_col = None
            return_rate_col = None
            smoothing_col = None
            
            for col_name, col_idx in headers.items():
                if col_name == '2024 AV Link':
                    column_mapping['2024 AV Link'] = col_idx
                elif col_name == 'AV Date':
                    column_mapping['AV Date'] = col_idx
                    av_date_col = col_idx
                elif col_name == 'Inflation Rate':
                    column_mapping['Inflation Rate'] = col_idx
                    inflation_col = col_idx
                elif col_name == 'Rate of Return on Pension Investments':
                    column_mapping['Rate of Return on Pension Investments'] = col_idx
                    return_rate_col = col_idx
                elif col_name == 'Smoothing':
                    column_mapping['Smoothing'] = col_idx
                    smoothing_col = col_idx
            
            # Find Reference columns
            if av_date_col:
                next_col_header = ws.cell(1, av_date_col + 1).value
                if next_col_header and ('Reference' in str(next_col_header) or next_col_header == 'Reference'):
                    column_mapping['AV Date Reference'] = av_date_col + 1
            
            if inflation_col:
                next_col_header = ws.cell(1, inflation_col + 1).value
                if next_col_header and ('Reference' in str(next_col_header) or next_col_header == 'Reference'):
                    column_mapping['Inflation Rate Reference'] = inflation_col + 1
            
            if return_rate_col:
                next_col_header = ws.cell(1, return_rate_col + 1).value
                if next_col_header and ('Reference' in str(next_col_header) or next_col_header == 'Reference'):
                    column_mapping['Rate of Return Reference'] = return_rate_col + 1
            
            if smoothing_col:
                next_col_header = ws.cell(1, smoothing_col + 1).value
                if next_col_header and ('Reference' in str(next_col_header) or next_col_header == 'Reference'):
                    column_mapping['Smoothing Reference'] = smoothing_col + 1
            
            # Check if required column exists
            if column_mapping['2024 AV Link'] is None:
                raise Exception("'2024 AV Link' column not found")
            
            # Track matched JSON entries
            matched_json_files = set()
            unmatched_json_files = {}
            
            # Process each row
            updates_count = 0
            for row_idx in range(2, ws.max_row + 1):
                av_link_cell = ws.cell(row_idx, column_mapping['2024 AV Link'])
                av_link_value = av_link_cell.value
                
                if not av_link_value:
                    continue
                
                # Extract filename and keep original case
                filename = self.extract_filename_from_path(str(av_link_value))
                filename_lower = filename.lower()
                
                # Find matching JSON entry using case-insensitive comparison
                matched_json_key = None
                
                # STEP 1: Try exact match first
                for json_key in json_data.keys():
                    if json_key.lower() == filename_lower:
                        matched_json_key = json_key
                        break
                
                # STEP 2: If no exact match found, try partial match (substring)
                if not matched_json_key:
                    for json_key in json_data.keys():
                        json_key_lower = json_key.lower()
                        if json_key_lower in filename_lower:
                            matched_json_key = json_key
                            break
                
                if matched_json_key:
                    data = json_data[matched_json_key]
                    matched_json_files.add(matched_json_key)
                    
                    # Update AV Date and its Reference
                    if column_mapping['AV Date'] and 'av_date' in data:
                        ws.cell(row_idx, column_mapping['AV Date']).value = data['av_date'].get('value', 'not sure')
                        if column_mapping['AV Date Reference']:
                            ws.cell(row_idx, column_mapping['AV Date Reference']).value = data['av_date'].get('document_page', 'not sure')
                    
                    # Update Inflation Rate and its Reference
                    if column_mapping['Inflation Rate'] and 'actuarial_inflation_rate' in data:
                        ws.cell(row_idx, column_mapping['Inflation Rate']).value = data['actuarial_inflation_rate'].get('value', 'not sure')
                        if column_mapping['Inflation Rate Reference']:
                            ws.cell(row_idx, column_mapping['Inflation Rate Reference']).value = data['actuarial_inflation_rate'].get('document_page', 'not sure')
                    
                    # Update Rate of Return and its Reference
                    if column_mapping['Rate of Return on Pension Investments'] and 'actuarial_return_rate' in data:
                        ws.cell(row_idx, column_mapping['Rate of Return on Pension Investments']).value = data['actuarial_return_rate'].get('value', 'not sure')
                        if column_mapping['Rate of Return Reference']:
                            ws.cell(row_idx, column_mapping['Rate of Return Reference']).value = data['actuarial_return_rate'].get('document_page', 'not sure')
                    
                    # Update Smoothing and its Reference
                    if column_mapping['Smoothing'] and 'smoothing_years' in data:
                        ws.cell(row_idx, column_mapping['Smoothing']).value = data['smoothing_years'].get('value', 'not sure')
                        if column_mapping['Smoothing Reference']:
                            ws.cell(row_idx, column_mapping['Smoothing Reference']).value = data['smoothing_years'].get('document_page', 'not sure')
                    
                    updates_count += 1
            
            # Find unmatched JSON entries
            for json_key, json_value in json_data.items():
                if json_key not in matched_json_files:
                    unmatched_json_files[json_key] = json_value
            
            # Save unmatched entries to JSON file
            if unmatched_json_files:
                unmatched_path = self.json_path.replace('.json', '_unmatched.json')
                with open(unmatched_path, 'w', encoding='utf-8') as f:
                    json.dump(unmatched_json_files, f, indent=2, ensure_ascii=False)
                print(f"Saved {len(unmatched_json_files)} unmatched entries to: {unmatched_path}")
            
            # Save workbook
            wb.save(self.excel_path)
            return updates_count, len(unmatched_json_files)
            
        except Exception as e:
            raise Exception(f"Error merging data: {str(e)}")
        
class ProcessingThread(QThread):
    """Thread to handle PDF processing in background."""
    
    progress_update = pyqtSignal(str)
    progress_value = pyqtSignal(int)
    finished = pyqtSignal(bool, str)
    
    def __init__(self, pdf_folder, json_filename, api_key, extraction_prompt_path):
        super().__init__()
        self.pdf_folder = pdf_folder
        self.json_filename = json_filename
        self.api_key = api_key
        self.extraction_prompt_path = extraction_prompt_path
        self.keywords = [
            "inflation", "return rate", "rate of return", "rate return",
            "returning rate", "rate of returning", "smoothing", "smooth",
            "investment return", "return on investment", "return of investment"
        ]
        self.consecutive_errors = 0  # Track consecutive errors
        self.max_consecutive_errors = 3  # Maximum allowed consecutive errors

    
    def run(self):
        try:
            # Step 1: Clean PDFs
            self.progress_update.emit("Step 1/2: Cleaning PDFs...")
            
            folder_name = os.path.basename(self.pdf_folder)
            cleaned_folder = os.path.join(os.path.dirname(self.pdf_folder), f"Cleaned {folder_name}")
            Path(cleaned_folder).mkdir(parents=True, exist_ok=True)
            
            pdf_files = [f for f in os.listdir(self.pdf_folder) if f.lower().endswith('.pdf')]
            
            if not pdf_files:
                self.finished.emit(False, "No PDF files found in selected folder")
                return
            
            cleaner = PDFCleaner(self.keywords) 

            output_json_path = os.path.join(os.path.dirname(self.pdf_folder), self.json_filename)

            # Load existing results
            if os.path.exists(output_json_path):
                with open(output_json_path, 'r', encoding='utf-8') as f:
                    results = json.load(f)
                self.progress_update.emit(f"Loaded existing results from {self.json_filename}")
            else:
                results = {}

            
            for i, pdf_file in enumerate(pdf_files):
                input_path = os.path.join(self.pdf_folder, pdf_file)
                output_path = os.path.join(cleaned_folder, pdf_file)
                
                # Check if cleaned PDF already exists
                if os.path.exists(output_path):
                    self.progress_update.emit(f"Skipping cleaning ({i+1}/{len(pdf_files)}): {pdf_file} (already cleaned)")
                    self.progress_value.emit(int((i + 1) / len(pdf_files) * 50))
                    continue

                self.progress_update.emit(f"Cleaning ({i+1}/{len(pdf_files)}): {pdf_file}")
                
                try:
                    new_pages, old_pages = cleaner.clean_pdf(input_path, output_path)
                    self.progress_update.emit(f"  ✓ Cleaned: {old_pages} → {new_pages} pages")
                except Exception as e:
                    self.progress_update.emit(f"  ✗ Error: {str(e)}")
                
                self.progress_value.emit(int((i + 1) / len(pdf_files) * 50))
            
            # Step 2: Extract data
            self.progress_update.emit("\nStep 2/2: Extracting data from cleaned PDFs...")
            
            
            
            extractor = DataExtractor(self.api_key, self.extraction_prompt_path)
            
            cleaned_pdf_files = [f for f in os.listdir(cleaned_folder) if f.lower().endswith('.pdf')]
            
            for i, pdf_file in enumerate(cleaned_pdf_files):
                # Check if we've hit the consecutive error limit
                if self.consecutive_errors >= self.max_consecutive_errors:
                    error_msg = f"Stopping process: {self.max_consecutive_errors} consecutive API errors encountered"
                    self.progress_update.emit(f"\n✗ {error_msg}")
                    
                    # Save results before stopping
                    with open(output_json_path, 'w', encoding='utf-8') as f:
                        json.dump(results, f, indent=2, ensure_ascii=False)
                    
                    self.finished.emit(False, error_msg)
                    return
                
                # Skip if already extracted
                if pdf_file in results and "error" not in results[pdf_file]:
                    self.progress_update.emit(f"Skipping ({i+1}/{len(cleaned_pdf_files)}): {pdf_file} (already extracted)")
                    continue
                
                pdf_path = os.path.join(cleaned_folder, pdf_file)
                self.progress_update.emit(f"Extracting ({i+1}/{len(cleaned_pdf_files)}): {pdf_file}")
                
                try:
                    extracted_data = extractor.extract_data_from_pdf(pdf_path)
                    
                    if extracted_data:
                        results[pdf_file] = extracted_data
                        self.progress_update.emit(f"  ✓ Successfully extracted data")
                        self.consecutive_errors = 0  # Reset counter on success
                    else:
                        raise Exception("Failed to extract data - returned None")
                        
                except Exception as e:
                    self.consecutive_errors += 1
                    error_msg = str(e)
                    self.progress_update.emit(f"  ✗ API Error ({self.consecutive_errors}/{self.max_consecutive_errors}): {error_msg}")
                    
                    results[pdf_file] = {
                        "error": f"Failed to extract data: {error_msg}",
                        "av_date": {"value": "not sure", "document_page": "not sure", "actual_pdf_page": "not sure"},
                        "actuarial_return_rate": {"value": "not sure", "document_page": "not sure", "actual_pdf_page": "not sure"},
                        "smoothing_years": {"value": "not sure", "document_page": "not sure", "actual_pdf_page": "not sure"},
                        "actuarial_inflation_rate": {"value": "not sure", "document_page": "not sure", "actual_pdf_page": "not sure"}
                    }
                
                # Save results after each PDF
                with open(output_json_path, 'w', encoding='utf-8') as f:
                    json.dump(results, f, indent=2, ensure_ascii=False)
                
                self.progress_value.emit(50 + int((i + 1) / len(cleaned_pdf_files) * 50))
                time.sleep(1)  # Rate limiting
            
            self.progress_update.emit(f"\n✓ Processing complete! Results saved to: {output_json_path}")
            self.finished.emit(True, output_json_path)
            
        except Exception as e:
            self.finished.emit(False, str(e))


class MergeThread(QThread):
    """Thread to handle Excel merging in background."""
    
    progress_update = pyqtSignal(str)
    finished = pyqtSignal(bool, str)
    
    def __init__(self, excel_path, json_path):
        super().__init__()
        self.excel_path = excel_path
        self.json_path = json_path
    
    def run(self):
        try:
            self.progress_update.emit("Merging JSON data to Excel file...")
            merger = ExcelMerger(self.excel_path, self.json_path)
            updates_count = merger.merge_data()
            self.progress_update.emit(f"\n✓ Successfully updated {updates_count} rows in Excel file")
            self.finished.emit(True, f"Updated {updates_count} rows")
        except Exception as e:
            self.finished.emit(False, str(e))


class MainWindow(QMainWindow):
    """Main GUI window."""
    
    def __init__(self):
        super().__init__()
        self.pdf_folder = ""
        self.excel_file = ""
        self.json_file = ""


        # self.api_key = "AIzaSyAwrkHTjjo_PJ5tQU9o-f0M_DJgC8mh5Zc"

        self.api_key = API_KEY

        self.extraction_prompt_path = "Extraction_Prompt Updated.txt"
        
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("PDF Data Extractor & Excel Merger")
        self.setGeometry(100, 100, 800, 650)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        layout = QVBoxLayout()
        
        # PDF Folder Selection
        layout.addWidget(QLabel("Step 1: Select PDF Folder"))
        
        pdf_layout = QHBoxLayout()
        self.pdf_label = QLabel("No folder selected")
        self.pdf_btn = QPushButton("Browse PDF Folder")
        self.pdf_btn.clicked.connect(self.browse_pdf_folder)
        pdf_layout.addWidget(self.pdf_label)
        pdf_layout.addWidget(self.pdf_btn)
        layout.addLayout(pdf_layout)
        
        # JSON Filename
        layout.addWidget(QLabel("\nStep 2: Set JSON Output Filename"))
        
        json_layout = QHBoxLayout()
        json_layout.addWidget(QLabel("Filename:"))
        self.json_input = QLineEdit("extracted_datapoints.json")
        json_layout.addWidget(self.json_input)
        layout.addLayout(json_layout)
        
        # Process Button
        self.process_btn = QPushButton("Process PDFs & Extract Data")
        self.process_btn.clicked.connect(self.process_pdfs)
        self.process_btn.setEnabled(False)
        layout.addWidget(self.process_btn)
        
        # Progress Bar
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)
        
        # Log Area
        layout.addWidget(QLabel("\nProcessing Log:"))
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        layout.addWidget(self.log_text)
        
        # Excel Merge Section
        layout.addWidget(QLabel("\n" + "="*50))
        layout.addWidget(QLabel("Step 3: Merge Data to Excel"))
        
        # JSON File Selection
        json_select_layout = QHBoxLayout()
        self.json_file_label = QLabel("No JSON file selected")
        self.json_file_btn = QPushButton("Browse JSON File")
        self.json_file_btn.clicked.connect(self.browse_json_file)
        json_select_layout.addWidget(self.json_file_label)
        json_select_layout.addWidget(self.json_file_btn)
        layout.addLayout(json_select_layout)
        
        # Excel File Selection
        excel_layout = QHBoxLayout()
        self.excel_label = QLabel("No Excel file selected")
        self.excel_btn = QPushButton("Browse Excel File")
        self.excel_btn.clicked.connect(self.browse_excel_file)
        excel_layout.addWidget(self.excel_label)
        excel_layout.addWidget(self.excel_btn)
        layout.addLayout(excel_layout)
        
        # Merge Button
        self.merge_btn = QPushButton("Merge JSON Data to Excel")
        self.merge_btn.clicked.connect(self.merge_to_excel)
        self.merge_btn.setEnabled(False)
        layout.addWidget(self.merge_btn)
        
        central_widget.setLayout(layout)
    
    def browse_pdf_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select PDF Folder")
        if folder:
            self.pdf_folder = folder
            self.pdf_label.setText(os.path.basename(folder))
            self.process_btn.setEnabled(True)
            self.log_text.append(f"Selected PDF folder: {folder}\n")
    
    def browse_json_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "Select JSON File", "", "JSON Files (*.json)")
        if file:
            self.json_file = file
            self.json_file_label.setText(os.path.basename(file))
            self.check_merge_button()
            self.log_text.append(f"Selected JSON file: {file}\n")
    
    def browse_excel_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx *.xls)")
        if file:
            self.excel_file = file
            self.excel_label.setText(os.path.basename(file))
            self.check_merge_button()
            self.log_text.append(f"Selected Excel file: {file}\n")
    
    def check_merge_button(self):
        if self.excel_file and self.json_file:
            self.merge_btn.setEnabled(True)
    
    def process_pdfs(self):
        if not self.pdf_folder:
            QMessageBox.warning(self, "Warning", "Please select a PDF folder first!")
            return
        
        # Check if extraction_prompt.txt exists
        if not os.path.exists(self.extraction_prompt_path):
            QMessageBox.warning(self, "Warning", f"'{self.extraction_prompt_path}' not found!")
            return
        
        self.process_btn.setEnabled(False)
        self.pdf_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        
        json_filename = self.json_input.text() or "extracted_datapoints.json"
        
        self.thread = ProcessingThread(self.pdf_folder, json_filename, self.api_key, self.extraction_prompt_path)
        self.thread.progress_update.connect(self.update_log)
        self.thread.progress_value.connect(self.progress_bar.setValue)
        self.thread.finished.connect(self.processing_finished)
        self.thread.start()
    
    def processing_finished(self, success, message):
        self.process_btn.setEnabled(True)
        self.pdf_btn.setEnabled(True)
        
        if success:
            self.json_file = message  # Store JSON file path
            self.json_file_label.setText(os.path.basename(message))
            self.check_merge_button()
            QMessageBox.information(self, "Success", "Processing completed successfully!")
        else:
            QMessageBox.critical(self, "Error", f"Processing failed: {message}")
    
    def merge_to_excel(self):
        if not self.excel_file or not self.json_file:
            QMessageBox.warning(self, "Warning", "Please select both JSON file and Excel file first!")
            return
        
        self.merge_btn.setEnabled(False)
        self.excel_btn.setEnabled(False)
        self.json_file_btn.setEnabled(False)
        self.log_text.append("\n" + "="*50 + "\n")
        
        self.merge_thread = MergeThread(self.excel_file, self.json_file)
        self.merge_thread.progress_update.connect(self.update_log)
        self.merge_thread.finished.connect(self.merge_finished)
        self.merge_thread.start()
    
    def merge_finished(self, success, message):
        self.merge_btn.setEnabled(True)
        self.excel_btn.setEnabled(True)
        self.json_file_btn.setEnabled(True)
        
        if success:
            QMessageBox.information(self, "Success", f"Merge completed! {message}")
        else:
            QMessageBox.critical(self, "Error", f"Merge failed: {message}")
    
    def update_log(self, message):
        self.log_text.append(message)
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )


def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()