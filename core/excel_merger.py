"""
core/excel_merger.py
Handles merging extracted JSON data into the Excel spreadsheet.
"""
import os
import json
import re
import shutil
import tempfile
import zipfile
import openpyxl


class ExcelMerger:
    """Merges JSON extraction data into a specified Excel workbook sheet."""

    SHEET_NAMES = ["AV Documents_Clean", "AV Data"]

    COLUMN_MAPPING_KEYS = [
        '2024 AV Link', '2024 AV Reference Document',
        'AV Date', 'AV Date Reference',
        'Inflation Rate', 'Inflation Rate Reference',
        'Rate of Return on Pension Investments', 'Rate of Return Reference',
        'Smoothing', 'Smoothing Reference'
    ]

    def __init__(self, excel_path, json_path):
        self.excel_path = excel_path
        self.json_path = json_path

    def extract_filename_from_path(self, path_string):
        if not path_string:
            return ""
        if '\\' in path_string or '/' in path_string:
            return os.path.basename(path_string)
        return path_string

    @staticmethod
    def _clean(value):
        """Return None if value is missing or the sentinel string 'not sure'."""
        if value is None:
            return None
        if str(value).strip().lower() == 'not sure':
            return None
        return value

    # ------------------------------------------------------------------
    # Comment-preservation helpers
    # ------------------------------------------------------------------

    def _save_preserving_comments(self, wb, save_path: str, active_sheet_name: str) -> None:
        """
        Save the workbook while reliably preserving 100% of the original comments, 
        threaded comments, VBA, macros, images, and other advanced XML tags 
        that openpyxl drops. Uses the original excel file as the base structure.
        """
        tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(tmp_fd)
        
        try:
            # 1. Save openpyxl modifications to a temporary zip
            wb.save(tmp_path)
            
            # 2. Prepare the final output path
            out_fd, out_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(out_fd)
            
            try:
                with zipfile.ZipFile(self.excel_path, 'r') as zin, \
                     zipfile.ZipFile(tmp_path, 'r') as zmod, \
                     zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                    
                    # We will use original structure as the base, replacing ONLY the active worksheet and sharedStrings.
                    orig_names = set(zin.namelist())
                    mod_names = set(zmod.namelist())
                    
                    written = set()
                    
                    # Discover the XML filename for the successfully modified `active_sheet_name`
                    # 1. Read workbook.xml to find the rId for the active sheet name
                    target_sheet_xml = None
                    if 'xl/workbook.xml' in orig_names and 'xl/_rels/workbook.xml.rels' in orig_names:
                        wb_content = zin.read('xl/workbook.xml').decode('utf-8')
                        escaped_name = active_sheet_name.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        
                        sheet_match = re.search(r'<sheet [^>]*name=["\']' + re.escape(escaped_name) + r'["\'][^>]*r:id=["\']([^"\']+)["\']', wb_content, re.IGNORECASE)
                        if sheet_match:
                            rid = sheet_match.group(1)
                            # 2. Read workbook.xml.rels to map rId to the actual xl/worksheets/sheetX.xml target
                            rels_content = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
                            rel_match = re.search(r'<Relationship [^>]*Id=["\']' + re.escape(rid) + r'["\'][^>]*Target=["\']([^"\']+)["\']', rels_content, re.IGNORECASE)
                            if rel_match:
                                target = rel_match.group(1) # e.g. 'worksheets/sheet2.xml'
                                target_sheet_xml = f"xl/{target}"
                    
                    # Read original sheets to rescue tags openpyxl drops (only needed for our target)
                    orig_sheets = {}
                    if target_sheet_xml and target_sheet_xml in orig_names:
                        orig_sheets[target_sheet_xml] = zin.read(target_sheet_xml).decode('utf-8')
                            
                    # A. First, write the modified active sheet, sharedStrings, and styles
                    for name in mod_names:
                        if name in ('xl/sharedStrings.xml', 'xl/styles.xml') or name == target_sheet_xml:
                            mod_content = zmod.read(name).decode('utf-8')
                            
                            # If it's a worksheet, rescue the advanced tags from the original
                            if name in orig_sheets:
                                orig_content = orig_sheets[name]
                                
                                # Rescue namespaces to prevent "Undeclared prefix" XML errors
                                orig_root_match = re.search(r'<worksheet([^>]*)>', orig_content, re.IGNORECASE)
                                mod_root_match = re.search(r'<worksheet([^>]*)>', mod_content, re.IGNORECASE)
                                
                                if orig_root_match and mod_root_match:
                                    orig_attrs = orig_root_match.group(1)
                                    mod_attrs = mod_root_match.group(1)
                                    
                                    # Find all xmlns definitions in the original root
                                    xmlns_matches = re.finditer(r'(xmlns:?[a-zA-Z0-9_-]*="[^"]*")', orig_attrs)
                                    for xml_match in xmlns_matches:
                                        full_attr = xml_match.group(1)
                                        # Extract just the prefix part
                                        prefix = full_attr.split('=')[0].strip()
                                        if prefix not in mod_attrs:
                                            mod_attrs += f' {full_attr}'
                                            
                                    mod_content = mod_content.replace(mod_root_match.group(0), f'<worksheet{mod_attrs}>')

                                # Tags to rescue from the original sheet in correct XML sequence
                                tags_to_rescue = [
                                    r'<drawing[^>]+>',
                                    r'<legacyDrawing[^>]+>',
                                    r'<legacyDrawingHF[^>]+>',
                                    r'<picture[^>]+>',
                                    r'<oleObjects[^>]*>.*?</oleObjects>',
                                    r'<controls[^>]*>.*?</controls>',
                                    r'<tableParts[^>]*>.*?</tableParts>',
                                    r'<extLst[^>]*>.*?</extLst>'
                                ]
                                
                                for pattern in tags_to_rescue:
                                    orig_match = re.search(pattern, orig_content, re.IGNORECASE | re.DOTALL)
                                    if orig_match:
                                        tag_html = orig_match.group(0)
                                        # Strip it from mod_content if openpyxl generated a corrupted/incomplete one
                                        mod_content = re.sub(pattern, '', mod_content, flags=re.IGNORECASE | re.DOTALL)
                                        # Append the pristine original tag
                                        mod_content = mod_content.replace('</worksheet>', f'{tag_html}</worksheet>')
                                        
                            zout.writestr(name, mod_content.encode('utf-8'))
                            written.add(name)
                            
                    # B. Check if sharedStrings was newly created by openpyxl
                    if 'xl/sharedStrings.xml' in written and 'xl/sharedStrings.xml' not in orig_names:
                        # We must inject its Override into [Content_Types].xml
                        if '[Content_Types].xml' in orig_names:
                            ct_content = zin.read('[Content_Types].xml').decode('utf-8')
                            if 'sharedStrings.xml' not in ct_content:
                                override = '<Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
                                ct_content = ct_content.replace('</Types>', f'{override}</Types>')
                                zout.writestr('[Content_Types].xml', ct_content.encode('utf-8'))
                                written.add('[Content_Types].xml')
                        
                        # We must also inject it into workbook.xml.rels
                        rels_path = 'xl/_rels/workbook.xml.rels'
                        if rels_path in orig_names:
                            rels_content = zin.read(rels_path).decode('utf-8')
                            if 'sharedStrings.xml' not in rels_content:
                                rel_id = f"rId{len(re.findall('<Relationship', rels_content)) + 1}"
                                rel = f'<Relationship Id="{rel_id}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" Target="sharedStrings.xml"/>'
                                rels_content = rels_content.replace('</Relationships>', f'{rel}</Relationships>')
                                zout.writestr(rels_path, rels_content.encode('utf-8'))
                                written.add(rels_path)

                    # C. Write EVERYTHING else from the original zip unchanged
                    for item in zin.infolist():
                        name = item.filename
                        if name not in written:
                            zout.writestr(item, zin.read(name))
                            written.add(name)
                            
                shutil.move(out_path, save_path)
            except Exception:
                if os.path.exists(out_path):
                    os.remove(out_path)
                raise
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    # ------------------------------------------------------------------

    def merge_data(self):
        """
        Merge JSON data into Excel file.
        Returns (updates_count, unmatched_count).
        """
        with open(self.json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)

        wb = openpyxl.load_workbook(self.excel_path)

        ws = None
        for sheet_name in self.SHEET_NAMES:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                break

        if ws is None:
            raise Exception(
                f"Neither '{self.SHEET_NAMES[0]}' nor '{self.SHEET_NAMES[1]}' "
                f"sheet was found in Excel file"
            )

        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[cell.value] = col_idx

        column_mapping = {k: None for k in self.COLUMN_MAPPING_KEYS}

        av_date_col = inflation_col = return_rate_col = smoothing_col = None

        for col_name, col_idx in headers.items():
            if col_name == '2024 AV Link':
                column_mapping['2024 AV Link'] = col_idx
            elif col_name == '2024 AV Reference Document':
                column_mapping['2024 AV Reference Document'] = col_idx
            elif col_name == 'AV Date':
                column_mapping['AV Date'] = col_idx
                av_date_col = col_idx
            elif col_name == 'Inflation Rate':
                column_mapping['Inflation Rate'] = col_idx
                inflation_col = col_idx
            elif col_name == 'Rate of Return on Pension Investments':
                column_mapping['Rate of Return on Pension Investments'] = col_idx
                return_rate_col = col_idx
            elif col_name == 'Smoothing':
                column_mapping['Smoothing'] = col_idx
                smoothing_col = col_idx

        # Resolve the AV link column: prefer '2024 AV Link', fall back to '2024 AV Reference Document'
        av_link_col = column_mapping['2024 AV Link'] or column_mapping['2024 AV Reference Document']

        def _try_ref(base_col, ref_key):
            """Detect the Reference column for a given data column.

            Handles two header layouts:
              Format 1 (with CL Note):  data | CL Note | Reference
              Format 2 (no CL Note):   data | Reference
            """
            if not base_col:
                return
            # Format 2: Reference immediately after the data column
            offset1_val = ws.cell(1, base_col + 1).value
            if offset1_val and 'Reference' in str(offset1_val):
                column_mapping[ref_key] = base_col + 1
                return
            # Format 1: CL Note sits between data and Reference
            offset1_is_cl_note = offset1_val and 'CL Note' in str(offset1_val)
            if offset1_is_cl_note:
                offset2_val = ws.cell(1, base_col + 2).value
                if offset2_val and 'Reference' in str(offset2_val):
                    column_mapping[ref_key] = base_col + 2

        _try_ref(av_date_col, 'AV Date Reference')
        _try_ref(inflation_col, 'Inflation Rate Reference')
        _try_ref(return_rate_col, 'Rate of Return Reference')
        _try_ref(smoothing_col, 'Smoothing Reference')

        if av_link_col is None:
            raise Exception(
                "Neither '2024 AV Link' nor '2024 AV Reference Document' "
                "column was found in Excel sheet"
            )

        matched_json_files = set()
        updates_count = 0

        for row_idx in range(2, ws.max_row + 1):
            av_link_cell = ws.cell(row_idx, av_link_col)
            av_link_value = av_link_cell.value
            if not av_link_value:
                continue

            filename = self.extract_filename_from_path(str(av_link_value))
            filename_lower = filename.lower()

            matched_json_key = None
            for json_key in json_data.keys():
                if json_key.lower() == filename_lower:
                    matched_json_key = json_key
                    break

            if not matched_json_key:
                for json_key in json_data.keys():
                    if json_key.lower() in filename_lower:
                        matched_json_key = json_key
                        break

            if matched_json_key:
                data = json_data[matched_json_key]
                matched_json_files.add(matched_json_key)

                if column_mapping['AV Date'] and 'av_date' in data:
                    ws.cell(row_idx, column_mapping['AV Date']).value = self._clean(data['av_date'].get('value'))
                    if column_mapping['AV Date Reference']:
                        ws.cell(row_idx, column_mapping['AV Date Reference']).value = self._clean(data['av_date'].get('document_page'))

                if column_mapping['Inflation Rate'] and 'actuarial_inflation_rate' in data:
                    ws.cell(row_idx, column_mapping['Inflation Rate']).value = self._clean(data['actuarial_inflation_rate'].get('value'))
                    if column_mapping['Inflation Rate Reference']:
                        ws.cell(row_idx, column_mapping['Inflation Rate Reference']).value = self._clean(data['actuarial_inflation_rate'].get('document_page'))

                if column_mapping['Rate of Return on Pension Investments'] and 'actuarial_return_rate' in data:
                    ws.cell(row_idx, column_mapping['Rate of Return on Pension Investments']).value = self._clean(data['actuarial_return_rate'].get('value'))
                    if column_mapping['Rate of Return Reference']:
                        ws.cell(row_idx, column_mapping['Rate of Return Reference']).value = self._clean(data['actuarial_return_rate'].get('document_page'))

                if column_mapping['Smoothing'] and 'smoothing_years' in data:
                    ws.cell(row_idx, column_mapping['Smoothing']).value = self._clean(data['smoothing_years'].get('value'))
                    if column_mapping['Smoothing Reference']:
                        ws.cell(row_idx, column_mapping['Smoothing Reference']).value = self._clean(data['smoothing_years'].get('document_page'))

                updates_count += 1

        unmatched = {k: v for k, v in json_data.items() if k not in matched_json_files}

        if unmatched:
            unmatched_path = self.json_path.replace('.json', '_unmatched.json')
            with open(unmatched_path, 'w', encoding='utf-8') as f:
                json.dump(unmatched, f, indent=2, ensure_ascii=False)

        self._save_preserving_comments(wb, self.excel_path, ws.title)
        return updates_count, len(unmatched)
